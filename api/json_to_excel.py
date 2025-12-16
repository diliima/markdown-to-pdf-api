import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import logging
from typing import Optional, List, Dict, Any

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class JsonToExcel:
    """
    Classe para converter dados JSON para planilhas Excel (.xlsx).
    Versão otimizada para ambiente serverless (Vercel).
    """
    
    def __init__(self):
        """
        Inicializa o conversor.
        """
        self.workbook = None
        
    def _setup_default_styles(self):
        """
        Define estilos padrão para a planilha.
        """
        # Estilo para cabeçalhos
        self.header_font = Font(
            name='Calibri',
            size=12,
            bold=True,
            color='FFFFFF'
        )
        
        self.header_fill = PatternFill(
            start_color='4472C4',
            end_color='4472C4',
            fill_type='solid'
        )
        
        self.header_alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        
        # Estilo para dados
        self.data_font = Font(
            name='Calibri',
            size=11
        )
        
        self.data_alignment = Alignment(
            horizontal='left',
            vertical='center'
        )
        
        # Bordas
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Estilo para linhas alternadas
        self.alt_fill = PatternFill(
            start_color='F2F2F2',
            end_color='F2F2F2',
            fill_type='solid'
        )
    
    def _apply_formatting(self, worksheet, data_rows: int):
        """
        Aplica formatação à planilha.
        
        Args:
            worksheet: Planilha do openpyxl
            data_rows: Número de linhas de dados
        """
        try:
            # Formatação do cabeçalho
            for cell in worksheet[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border
            
            # Formatação dos dados
            for row_idx in range(2, data_rows + 2):
                for cell in worksheet[row_idx]:
                    cell.font = self.data_font
                    cell.alignment = self.data_alignment
                    cell.border = self.border
                    
                    # Linhas alternadas
                    if row_idx % 2 == 0:
                        cell.fill = self.alt_fill
            
            # Ajusta largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Define largura mínima e máxima
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            # Se a formatação falhar, continua sem formatação
            logger.warning(f"Erro na formatação: {str(e)}")
    
    def json_to_excel_file(self, json_data: List[Dict[str, Any]], output_path: str, 
                          sheet_name: str = "Dados", apply_formatting: bool = True) -> bool:
        """
        Converte dados JSON para arquivo Excel.
        
        Args:
            json_data: Lista de dicionários com os dados
            output_path: Caminho onde salvar o arquivo Excel
            sheet_name: Nome da aba na planilha
            apply_formatting: Se deve aplicar formatação
            
        Returns:
            bool: True se a conversão foi bem-sucedida
        """
        try:
            logger.info("Iniciando conversão de JSON para Excel...")
            
            if not json_data or not isinstance(json_data, list):
                logger.error("Dados JSON devem ser uma lista não vazia")
                return False
            
            # Converte para DataFrame do pandas
            df = pd.DataFrame(json_data)
            
            # Cria workbook
            self.workbook = Workbook()
            worksheet = self.workbook.active
            worksheet.title = sheet_name[:31]  # Excel limita nomes de aba a 31 caracteres
            
            # Configura estilos se necessário
            if apply_formatting:
                self._setup_default_styles()
            
            # Adiciona dados à planilha
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
            
            # Aplica formatação se solicitada
            if apply_formatting:
                self._apply_formatting(worksheet, len(df))
            
            # Salva o arquivo
            self.workbook.save(output_path)
            
            logger.info(f"Planilha Excel gerada com sucesso: {output_path}")
            logger.info(f"Dados processados: {len(df)} linhas, {len(df.columns)} colunas")
            
            return True
            
        except Exception as e:
            logger.error(f"Erro ao converter JSON para Excel: {str(e)}")
            return False
    
    def validate_json_data(self, json_data: Any) -> tuple:
        """
        Valida se os dados JSON são adequados para conversão.
        
        Args:
            json_data: Dados a serem validados
            
        Returns:
            tuple: (is_valid, error_message)
        """
        try:
            # Verifica se é uma lista
            if not isinstance(json_data, list):
                return False, "Dados devem ser uma lista"
            
            # Verifica se não está vazio
            if not json_data:
                return False, "Lista de dados não pode estar vazia"
            
            # Verifica se todos os itens são dicionários
            for i, item in enumerate(json_data):
                if not isinstance(item, dict):
                    return False, f"Item {i} não é um dicionário válido"
            
            # Verifica se todos os dicionários têm chaves consistentes
            if json_data:
                first_keys = set(json_data[0].keys()) if json_data[0] else set()
                for i, item in enumerate(json_data[1:], 1):
                    current_keys = set(item.keys()) if item else set()
                    if current_keys != first_keys:
                        # Permite chaves diferentes se não for muito divergente
                        if len(current_keys.symmetric_difference(first_keys)) > len(first_keys) * 0.5:
                            return False, f"Item {i} tem chaves muito diferentes do primeiro item"
            
            return True, "Dados válidos"
            
        except Exception as e:
            return False, f"Erro na validação: {str(e)}"
    
    def json_string_to_excel(self, json_string: str, output_path: str, 
                           sheet_name: str = "Dados", apply_formatting: bool = True) -> bool:
        """
        Converte string JSON para arquivo Excel.
        
        Args:
            json_string: String contendo dados JSON
            output_path: Caminho onde salvar o arquivo Excel
            sheet_name: Nome da aba na planilha
            apply_formatting: Se deve aplicar formatação
            
        Returns:
            bool: True se a conversão foi bem-sucedida
        """
        try:
            # Parse da string JSON
            json_data = json.loads(json_string)
            
            # Valida os dados
            is_valid, error_msg = self.validate_json_data(json_data)
            if not is_valid:
                logger.error(f"Dados JSON inválidos: {error_msg}")
                return False
            
            # Converte para Excel
            return self.json_to_excel_file(json_data, output_path, sheet_name, apply_formatting)
            
        except json.JSONDecodeError as e:
            logger.error(f"Erro ao fazer parse do JSON: {str(e)}")
            return False
        except Exception as e:
            logger.error(f"Erro na conversão string JSON para Excel: {str(e)}")
            return False